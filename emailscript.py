import win32com.client
import win32com.client
import re
#other libraries to be used in this script
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta

# define the entry ID of the folder you want to access
folder_entry_id = '000000007419B56D85F6034CA70ADB117071924B01009DA216306737BE4792F0720A61CD4818000016643FE60000'

# create a MAPI session
outlook = win32com.client.Dispatch("Outlook.Application")
session = outlook.Session

# logon to the session
session.Logon()

try:
    # get the folder object using the entry ID
    folder = session.GetFolderFromID(folder_entry_id)
    # print("Folder name:", folder.Name)
    # print("Number of items:", folder.Items.Count)
    messages=folder.items
    messages.Sort("[ReceivedTime]", True)
    messages = messages.Restrict("[SenderEmailAddress] = 'do-not-reply@accelq.com'")
    message=messages.GetFirst()

    body_content=message.body
    result=re.search(r"\d\d: \d\d AM",body_content)
    print(result)
    starttime=result.group()
    starttime=starttime[0:2]
    print(starttime)
    while(starttime!="07"):
        message=messages.GetNext()
        body_content=message.body
        result=re.search(r"\d\d: \d\d AM",body_content)
        print(result)
        starttime=result.group()
        starttime=starttime[0:2]
        print(starttime)
    result=re.search(r"\d\d:\d\d:\d\d",body_content)
    jobid=message.subject[50:62]
    duration=result.group()

    print("Duration: "+duration)
    # print(received_dt)
    print(jobid)
    
    with open('log.txt',mode='a') as f:
        f.write("\n"+jobid+", "+duration)

except Exception as e:
    print("Error accessing folder:", str(e))

finally:
    # log off from the session
    session.Logoff()
    f.close()

# wb= load_workbook('Testrun_log.xlsx')
# ws=wb['Sheet1']
# ws=wb.active
# ws['A1']='Duration'
# print(ws['A1'].value)
